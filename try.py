#list1 = [1,2,3,4,4,5,5,5,5,5,6,7,7,8,8,8]
#i = 0
#while i < len(list1)-1:
#	count1 = list1.count(list1[i])
#	print(list1[i],count1)
#	i = i + count1
#	if i == len(list1)-1:
#		print(list1[i],list1.count(list1[i]))

from openpyxl import Workbook, load_workbook


#print(sheet.max_row)



#sheet.auto_filter.ref = "A1:B19"
#sheet.auto_filter.add_sort_condition("B2:B15")


import win32com.client

excel = win32com.client.Dispatch("Excel.Application")

wb = excel.Workbooks.Open(r"C:\Users\Adyasha\eclipse-workspace\Accelarate\PDFextract.xlsx")
ws = wb.Worksheets('Sheet')

ws.Range('A2:D21').Sort(Key1=ws.Range('D1'), Order1=1, Orientation=1)

wb.Save()
excel.Application.Quit()

workbook = load_workbook(filename=r"C:\Users\Adyasha\eclipse-workspace\Accelarate\PDFextract.xlsx") #Enter filename of keyword
sheet = workbook.active

i = 2

#while i <= sheet.max_row:
contents = []
keywords = []
for value in sheet.iter_rows(max_col=4, values_only=True):
	contents.append(value[3])
	keywords.append(value[1])

wb = Workbook()
ws = wb.active

ws['A1'] = 'Document Name'
ws['B1'] = 'Keyword'
ws['C1'] = 'Page No'
ws['D1'] = 'Content'

count2 = 1
print(contents)
while i <= sheet.max_row :
	count2 = count2 + 1
	count1 = 1
	count1 = contents.count(sheet['D' + str(i)].value)
	if count1 == 1:
		print(i, count2, sheet['D' + str(i)].value)
		ws['A' + str(count2)] = sheet['A' + str(i)].value
		ws['B' + str(count2)] = sheet['B' + str(i)].value
		ws['C' + str(count2)] = sheet['C' + str(i)].value
		ws['D' + str(count2)] = sheet['D' + str(i)].value
	else:
		print(i,count1,count2)
		str1 = ''
		for k in range(i, i+count1):
			str1 = str1 + sheet['B' + str(k)].value
			str1 = str1 + ','
		print(str1)
		ws['A' + str(count2)] = sheet['A' + str(i)].value
		ws['B' + str(count2)] = str1[:-1]
		ws['C' + str(count2)] = sheet['C' + str(i)].value
		ws['D' + str(count2)] = sheet['D' + str(i)].value
	i = i + count1
	print(i)
	if i == sheet.max_row:
		ws['A' + str(count2)] = sheet['A' + str(i)].value
		ws['B' + str(count2)] = sheet['B' + str(i)].value
		ws['C' + str(count2)] = sheet['C' + str(i)].value
		ws['D' + str(count2)] = sheet['D' + str(i)].value

	
wb.save("yashpal.xlsx")
